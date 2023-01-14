import time
from openpyxl import Workbook, load_workbook
import random

counter1 = 0
counter2 = 0
counter3 = 0
counter4 = 0

starttime = time.time()

wb = load_workbook("C:\\Users\\Ced\\Desktop\\ioo\\dict.xlsx")
sheet= wb.active 

#la colonne
column='A'

#range
lerange= [sheet[column + str(row)].value for row in range(1, sheet.max_row + 1)]

ques = random.choice(lerange)


#loop with time limit, when one player reaches 10 points, loop ends
while counter1 < 4 or counter2 < 4 or counter3 < 4 or counter4 < 4:
    
    ans1=input("Écrivez un mot commençant par les deux dernières lettres du mot : " + ques + "\n")
  
    for i in range(len(ans1)):
        if ans1[i:i+2] == ques[-2:] and ans1 in lerange:
            starttime = time.time()
            counter1 += 1
            ques=ans1
            print('Bonne réponse!')
            print(counter1, counter2, counter3, counter4)
            break
            
        else:
            starttime = time.time()
            print('Mauvaise réponse!')
            print(counter1, counter2, counter3, counter4)
            break
    ans2=input("Écrivez un mot commençant par les deux dernières lettres du mot : " + ques + "\n")
    for i in range(len(ans2)):
        if ans2[i:i+2] == ques[-2:] and ans2 in lerange:
            starttime = time.time()
            counter2 += 1
            ques=ans2
            print('Bonne réponse!')
            print(counter1, counter2, counter3, counter4)
            break
            
        else:
            starttime = time.time()
            print('Mauvaise réponse!')
            print(counter1, counter2, counter3, counter4)
            break
    ans3=input("Écrivez un mot commençant par les deux dernières lettres du mot : " + ques + "\n")
    for i in range(len(ans3)):
        if ans3[i:i+2] == ques[-2:] and ans3 in lerange:
            starttime = time.time()
            counter3 += 1
            ques=ans3
            print('Bonne réponse!')
            print(counter1, counter2, counter3, counter4)
            break
            
        else:
            starttime = time.time()
            print('Mauvaise réponse!')
            print(counter1, counter2, counter3, counter4)
            break
    ans4=input("Écrivez un mot commençant par les deux dernières lettres du mot : " + ques + "\n")
    for i in range(len(ans4)):
        if ans4[i:i+2] == ques[-2:] and ans4 in lerange:
            starttime = time.time()
            counter4 += 1
            ques=ans4
            print('Bonne réponse!')
            print(counter1, counter2, counter3, counter4)
            break
            
        else:
            starttime = time.time()
            print('Mauvaise réponse!')
            print(counter1, counter2, counter3, counter4)
            break
            

